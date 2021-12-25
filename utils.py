import io
from openpyxl import load_workbook

# Tương tác với file .txt
class File_Interact():
    def __init__(self, file_name):
        self.file_name = file_name
        
    def write_file(self, ndung):
        f = io.open(self.file_name, 'w', encoding='utf-8') # w thì sẽ xóa ndung cũ trên file và thay thế ndung mới
        f.write(ndung)
        f.close()
        
    def write_file_from_list(self, list_lines):
        f = io.open(self.file_name, 'w', encoding='utf-8')
        f.write('\n'.join(list_lines))
        f.close()
        
    def write_file_newline(self, ndung_line):
        f = io.open(self.file_name, 'a', encoding='utf-8') # a sẽ ghi thêm ndung vào file
        f.write('%s\n'%ndung_line)
        f.close()
        
    def read_file(self):
        f = io.open(self.file_name, 'r', encoding='utf-8')
        ndung = f.read()
        f.close()
        return ndung
    
    def read_file_list(self):
        f = io.open(self.file_name, 'r', encoding='utf-8')
        ndung = f.read()
        f.close()
        return ndung.split('\n')

# Tương tác với file .xlsx   
class Excel_Interact():
    def __init__(self, file_name):
        self.file_name = file_name
        
    def get_value_excel(self, sheet_name, cell_name):
        wb = load_workbook(filename=self.file_name)
        sheet_ranges = wb[sheet_name]
        wb.close()
        return sheet_ranges[cell_name].value
    
    def update_value_excel(self, sheet_name, cell_name, new_value):
        wb = load_workbook(filename=self.file_name)
        wb[sheet_name][cell_name].value = new_value
        wb.close()
        wb.save(self.file_name)

# Get proxies đang hoạt động ok và delete proxies bad đồng thời save proxies bad đó vào 1 file #.
class Proxies_Interact():
    def __init__(self, file_name):
        self.file_name = file_name
        
    def get_proxies_good(self):
        pass
    
    def delete_proxies_bad(self):
        pass
     
if __name__ == "__main__":
    pass
        